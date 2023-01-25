import openpyxl as op
import random
import os
import telebot
from telebot import types

file_name = 'my_artikel.xlsx'
new_words = []
bot = telebot.TeleBot('5875104312:AAECXNFjXGzcf6j1rrbs3P_qRQvBndVgxU8')
current_answer = 0
current_words = ''
current_article = ''
current_translate = ''
count_words = 0
rand_words = []
words_list_excel_3 = []


def read_list_from_excel(file_exel: str) -> list:
    '''создаем функцию по состалению из файла эксель списка'''

    words_list_local = []
    wb = op.load_workbook(file_exel, data_only=True)
    sheet = wb.active
    max_rows = sheet.max_row

    for i in range(2, max_rows + 1):
        article = sheet.cell(row=i, column=2).value
        words = sheet.cell(row=i, column=3).value
        translate = sheet.cell(row=i, column=4).value

        if words is not None:
            if len(words) != 0:
                words_list_local.append([article, words, translate])
    return words_list_local


def check_result_file(file_txt: str) -> list:
    '''создаем фунцию проверки наличие файла для записи результата'''
    new_words_local = []

    if not os.path.exists(file_txt):
        open(file_txt, 'tw', encoding='utf-8')
    else:
        with open(file_txt, 'r', encoding='utf-8') as f:
            new_words_local = f.read().splitlines()

    return new_words_local


# тело бота

@bot.message_handler(commands=['start'])
def start_message(message):
    print('run.start_message')
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton('Новичок')
    item2 = types.KeyboardButton('Серядничок')
    item3 = types.KeyboardButton('Мозжечок')
    markup.add(item1, item2, item3)
    msg = bot.send_message(message.chat.id, "Привет, {0.first_name}👋 !\n"
                                            "Я бот-тренер артиклей в немецком языке\n"
                                            "Выберете уровень для тренировки".format(message.from_user, bot.get_me()),
                           parse_mode='html', reply_markup=markup)
    bot.register_next_step_handler(message, main)


@bot.message_handler(content_types=['text'])
def main(message, words_list_excel=None):
    global count_words, current_article

    print(f'run.main {count_words} {current_article}')
    count_true_anwords = 0
    if message.chat.type == 'private':

        if message.text == 'Новичок':
            count_words = 10
        elif message.text == 'Серядничок':
            count_words = 30
        elif message.text == 'Мозжечок':
            count_words = 50
        else:
            count_words = 10

        if message.text in ['Новичок', 'Серядничок', 'Мозжечок']:
            # генерируем список слов
            global rand_words, words_list_excel_3

            words_list_excel_2 = read_list_from_excel(file_name)
            new_words = check_result_file('new_words.txt')

            # удаляем в списке для изучения слова из заученных
            words_list_excel = []
            for i in range(0, len(words_list_excel_2)):
                if not words_list_excel_2[i][1] in new_words:
                    words_list_excel.append(words_list_excel_2[i])

            for i in range(0, count_words):
                rand_words = random.randint(0, len(words_list_excel))
                words_list_excel_3.append(words_list_excel[rand_words])

        if message.text in ['der', 'die', 'das']:
            global current_answer
            if current_article == message.text:
                bot.send_message(message.chat.id, f'{" ".join(words_list_excel_3[current_answer])}\nвы угадали')
                count_true_anwords += 1
                new_words.append(words_list_excel_3[current_answer][1])
            else:
                bot.send_message(message.chat.id, f'{" ".join(words_list_excel_3[current_answer])}'
                                                  f'\nслово добавленов спирок повторения')
            current_answer += 1

            if count_words == current_answer:
                bot.send_message(message.chat.id,
                                 f'''количество правильных ответов {count_true_anwords} из {count_words}
                              Эффективность тренировки: {round((count_true_anwords / count_words * 100), 2)} процентов''')
                bot.send_message(message.chat.id,
                                 f'''результат освоения вашего курса {len(new_words)} из {len(words_list_excel_2)}
                               Статистика освония курса: {round((len(new_words) / len(words_list_excel_2) * 100), 2)} процентов''')
                # markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                # item1 = types.KeyboardButton("Закончить")
                # item2 = types.KeyboardButton("Еще разок")
                # markup.add(item1, item2)
                # if message.text == 'Закончить':
                #     bot.send_message(message.chat.id, 'Нажмите /start')
                # else:
                #     bot.send_message(message.chat.id, 'Пока. Tschüss')

        bot.send_message(message.chat.id, words_list_excel_3[current_answer][1])
        current_article = words_list_excel_3[current_answer][0]
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('der')
        item2 = types.KeyboardButton('die')
        item3 = types.KeyboardButton('das')
        markup.add(item1, item2, item3)
        msg = bot.send_message(message.chat.id, "выберете правильный артикль".format(message.from_user,
                                                                                     bot.get_me()),
                               parse_mode='html', reply_markup=markup)

        with open("new_words.txt", "w", encoding="utf-8") as f:
            for i in range(0, len(new_words)):
                f.write(new_words[i] + '\n')


if __name__ == '__main__':
    """
    main bot function
    """

    bot.skip_pending = True
    bot.polling(none_stop=True)
